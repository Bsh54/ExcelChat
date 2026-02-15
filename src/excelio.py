import fontTools.misc.cython
from openpyxl import load_workbook, workbook
from openpyxl.cell.cell import MergedCell
from src.observer import Observer
from src.memo import ModificationType
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
import os
import pandas as pd
from copy import copy


class ExcelAgent(Observer):
    def __init__(self):
        self.wb = None
        self.ws = None
        self.answer_ws = None
        self.plot_ws = None
        self.num_rows = 0
        self.num_cols = 0

    def load(self, filename):
        wb = load_workbook(filename=filename)
        if self.is_valid(wb):
            self.wb = wb
            self.ws = wb.active
            self.num_rows = self.ws.max_row
            self.num_cols = self.ws.max_column
            # Suppression de la création de la feuille "Réponse IA" ici
            return True
        return False

    def save(self, filename, fig_dir=None):
        if not self.is_opened():
            return
        if fig_dir is not None:
            self.insert_image(fig_dir)

        # S'assurer que la feuille "Réponse IA" est supprimée si elle existe
        if "Réponse IA" in self.wb.sheetnames:
            std = self.wb["Réponse IA"]
            self.wb.remove(std)

        self.wb.save(filename)

    def is_valid(self, wb):
        # On autorise désormais les cellules fusionnées pour gérer les fichiers complexes
        return True

    def is_opened(self):
        return self.wb is not None and self.ws is not None

    def _update(self, subject):
        if not self.is_opened():
            return
        modifications = subject.modification
        mtype = modifications.mtype
        item_infos = modifications.item_infos

        # Récupération de l'offset de l'en-tête depuis le sujet (EnhancedTable)
        header_offset = getattr(subject, 'header_row_idx', 1)

        if mtype == ModificationType.NEW_TABLE:
            # Traiter la mise à jour globale même si item_infos est vide
            df = modifications.df
            if df is None: return
            old_max_row = self.ws.max_row

            # On ne touche aux en-têtes que s'ils sont valides (pas Unnamed)
            for j, col_name in enumerate(df.columns):
                if "Unnamed" not in str(col_name):
                    self.ws.cell(row=header_offset, column=j+1, value=col_name)

            # Réécrire les données en conservant le style et sans tronquer le reste du fichier
            for i, row in enumerate(df.values):
                current_row = i + header_offset + 1
                for j, value in enumerate(row):
                    val_to_write = None if pd.isna(value) else value
                    cell = self.ws.cell(row=current_row, column=j+1)

                    # (Gestion du style déjà implémentée précédemment)
                    if current_row > old_max_row and current_row > header_offset + 1:
                        prev_cell = self.ws.cell(row=current_row - 1, column=j+1)
                        if prev_cell.has_style:
                            cell.font = copy(prev_cell.font)
                            cell.border = copy(prev_cell.border)
                            cell.fill = copy(prev_cell.fill)
                            cell.number_format = copy(prev_cell.number_format)
                            cell.alignment = copy(prev_cell.alignment)

                    cell.value = val_to_write

            # Ajuster la taille du fichier physique si nécessaire (Suppression des lignes en trop)
            new_max_row = len(df) + header_offset
            if old_max_row > new_max_row:
                # On supprime les lignes à partir de la fin pour ne pas décaler les index
                self.ws.delete_rows(new_max_row + 1, old_max_row - new_max_row)

            self.num_rows = self.ws.max_row
            self.num_cols = self.ws.max_column
            return

        if mtype == ModificationType.UPDATE_INPLACE:
            for item_info in item_infos:
                index = item_info.index
                text = item_info.text
                dtype = item_info.dtype
                try:
                    data = dtype(text)
                except:
                    data = float(text)
                # L'index est désormais absolu (0-based pour la feuille)
                # donc on utilise row_bias=0
                excel_cell_index = self.index_to_excel_index(*index, row_bias=0)
                self.ws[excel_cell_index] = data
                style = self.translate_style(item_info)
                self.apply_style(self.ws, excel_cell_index, style)
            return

        # Ajout du support pour les suppressions
        if mtype == ModificationType.DELETE_SCALAR:
            self.delete_scalar(item_infos)
        elif mtype == ModificationType.DELETE_ROW:
            self.delete_row(item_infos)
        elif mtype == ModificationType.DELETE_COLUMN:
            self.delete_column(item_infos)
        elif mtype == ModificationType.INSERT_ROW:
            self.insert_row(item_infos)
        elif mtype == ModificationType.INSERT_COLUMN:
            self.insert_column(item_infos)

    def insert_row(self, item_infos):
        for item_info in item_infos:
            index = item_info.index
            text = item_info.text
            dtype = item_info.dtype
            data = dtype(text)
            excel_cell_index = self.index_to_excel_index(*index)
            self.ws[excel_cell_index] = data
            style = self.translate_style(item_info)
            self.apply_style(self.ws, excel_cell_index, style)

    def insert_column(self, item_infos):
        for item_info in item_infos:
            index = item_info.index
            text = item_info.text
            dtype = item_info.dtype
            data = dtype(text)
            excel_cell_index = self.index_to_excel_index(*index)
            self.ws[excel_cell_index] = data
            style = self.translate_style(item_info)
            self.apply_style(self.ws, excel_cell_index, style)

    def insert_scalar(self, item_infos):
        item_info = item_infos[0]
        index = item_info.index
        text = item_info.text
        dtype = item_info.dtype
        data = dtype(text)
        excel_cell_index = self.index_to_excel_index(*index)
        self.ws[excel_cell_index] = data
        style = self.translate_style(item_info)
        self.apply_style(self.ws, excel_cell_index, style)

    def delete_scalar(self, item_infos):
        item_info = item_infos[0]
        row, column = item_info.index

        self.ws.cell(row=row, column=column, value=None)

    def delete_row(self, item_infos):
        item_info = item_infos[0]
        row, column = item_info.index
        # row est 0-based dans le tableau, donc row+1 dans Excel
        self.ws.delete_rows(row + 1)

    def delete_column(self, item_infos):
        item_info = item_infos[0]
        row, column = item_info.index
        # column est 0-based, donc column+1 dans Excel
        self.ws.delete_cols(column + 1)

    def update_inplace(self, item_infos):
        item_info = item_infos[0]
        index = item_info.index
        text = item_info.text
        dtype = item_info.dtype
        data = dtype(text)
        excel_cell_index = self.index_to_excel_index(*index)
        self.ws[excel_cell_index] = data
        style = self.translate_style(item_info)
        self.apply_style(self.ws, excel_cell_index, style)

    def index_to_excel_index(self, row, col, row_bias=0):
        excel_col = ''
        if col == 0:
            excel_col = 'A'
        else:
            col += 1
            while col > 0:
                remainder = (col - 1) % 26
                result = chr(ord('A') + remainder)
                excel_col = result + excel_col
                col = (col - 1) // 26

        excel_row = row + 1 + row_bias
        return f'{excel_col}{excel_row}'

    def translate_style(self, item_info):
        font = item_info.font
        font_size = item_info.font_size
        font_color = item_info.font_color.name()
        bg_color = item_info.bg_color.name()
        font_name = font.family()
        return font_name, font_size, 'FF' + font_color[1:], 'FF' + bg_color[1:]

    def apply_style(self, sheet, index, style):
        font_name, font_size, font_color, bg_color = style
        font = Font(name=font_name,
                    size=font_size,
                    color=font_color)
        fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color=bg_color)
        sheet[index].font = font
        sheet[index].fill = fill

    def insert_image(self, fig_dir):
        imgnames = os.listdir(fig_dir)
        if len(imgnames) == 0:
            return
        for i, imgname in enumerate(imgnames):
            plot_ws = self.wb.create_sheet("graphique{}".format(i+1))
            imgname = os.path.join(fig_dir, imgname)
            img = Image(imgname)
            plot_ws.add_image(img, 'A{}'.format(i + 1))
